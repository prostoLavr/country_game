from openpyxl import load_workbook
import os
import pickle
import argparse


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--file', type=str, default=None)
    parser.add_argument('-m', '--map', type=str, default=None)
    args = parser.parse_args()
    if args.file is None:
        read_file = input('Excel file: ').strip()
    else:
        read_file = args.file

    if not os.path.isfile(read_file):
        return 'File is not found'

    wb = load_workbook(read_file)
    ws = wb.active
    map_list = [[cell.value for cell in row] for row in ws.iter_rows()]
    map_list = [[('air' if cell is None else cell) for cell in row] for row in map_list]
    print(map_list)

    if not map_list:
        return 'File is empty' 
    if args.map is None:
        map_name = input('Map name: ').strip().replace('_', '-')
    else:
        map_name = args.map

    def formating(i, j):
        return os.path.join('res', 'compiled_maps', f'{map_name}_{i}_{j}.pickle')

    print(f'map shape: {(len(map_list), len(map_list[0]))}')
    for i in range(len(map_list) // 9):
        for j in range(len(map_list[0]) // 16):
            write_file = formating(i, j)
            map_to_save = [lst[(j * 16):((j + 1) * 16)] for lst in map_list[i * 9:(i + 1) * 9]]
            if all(map(lambda lst: None not in lst, map_to_save)) and len(map_to_save) and len(map_to_save[0]):
                with open(write_file, 'wb') as mapfile:
                    pickle.dump(map_to_save, mapfile)
                    print(f'File {write_file} saved')
    return ''


if __name__ == "__main__":
    while (res := main()) != '':
        print(f'\n{res}\nRETRY\n') 
    print('Success!')

